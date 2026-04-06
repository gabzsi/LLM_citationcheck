"""
━━━━━━━━━━━━━━━━━━━━━━━
Analyzes the Excel output from citation_lookup.py.

Classification (DOI NOT used):
  ✅ Valid   — paper found + title consistent → publisher bucket
  ❌ Invalid — paper not found or title mismatch

Invalidity comparison:
  Primary  — based on status + title_match only
  Extended — same as primary PLUS citations with a DOI mismatch (❌ doi_match)

Publisher Preference sheets compare two columns:
  (Val)  — 'publisher'          — validated publisher from Crossref/lookup
  (Orig) — 'journal_publisher'  — publisher extracted from the AI-generated citation

Sheet structure:
  1.  Invalidity_Comparison        — primary vs extended invalidity rates
  2.  All_Topics_Combined          — full dataset summary
  3.  <topic>                       — per-topic summary (full dataset)
  4.  NonInvalid_Combined          — same summaries on valid subset only
  5.  NonInvalid_<topic>           — per-topic on valid subset
  6.  Publication_Years_All        — year distribution, all citations
  7.  Publication_Years_Valid      — year distribution, valid only
  8.  Top_5_Journals_All           — top journals, all citations
  9.  Top_5_Journals_Valid         — top journals, valid only
  10. Publisher_Preference_All     — publisher comparison (all)
  11. Publisher_Preference_NonInvalid — publisher comparison (valid only)
"""

import re
import os
import html
import tkinter as tk
from tkinter import filedialog

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Column names ──────────────────────────────────────────────────────────────
COL_AI_ENGINE = 'ai_engine'
COL_TOPIC     = 'topic'
COL_YEAR      = 'year'
COL_JOURNAL   = 'journal'
COL_PUBLISHER = 'publisher'
COL_DOI       = 'doi'
COL_STATUS    = 'status'
COL_MATCH_SCORE = 'match_score'
COL_AUTHOR_M  = 'author_match'
COL_TITLE_M   = 'title_match'
COL_JOURNAL_M = 'journal_match'
COL_YEAR_M    = 'year_match'
COL_DOI_M     = 'doi_match'
COL_STRATEGY  = 'query_strategy'
COL_ORIG      = 'original_citation'
COL_FOUND     = 'found_citation'

COL_JOURNAL_PUB = 'journal_publisher'   # AI-extracted publisher (Orig column)

REQUIRED_COLUMNS = {
    COL_AI_ENGINE, COL_TOPIC, COL_YEAR, COL_JOURNAL,
    COL_PUBLISHER, COL_STATUS, COL_AUTHOR_M, COL_TITLE_M, COL_JOURNAL_M
}

# ── Publisher buckets ─────────────────────────────────────────────────────────
PUBLISHER_BUCKETS = [
    'Nature/Springer', 'Science/AAAS', 'ACS', 'RSC', 'Wiley', 'Elsevier',
    'Other', 'No Publisher'
]
INVALID_BUCKETS = ['Invalid']
TARGET_COLUMNS  = PUBLISHER_BUCKETS + INVALID_BUCKETS

# ── Styling ───────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
INVALID_FILL = PatternFill("solid", fgColor="FFC7CE")
NO_PUB_FILL  = PatternFill("solid", fgColor="E2EFDA")  # light green-grey for No Publisher
SUBHEAD_FILL = PatternFill("solid", fgColor="D9E1F2")
WARN_FILL    = PatternFill("solid", fgColor="FCE4D6")  # light orange for extended
WHITE_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BOLD_FONT    = Font(name="Arial", bold=True, size=10)
NORMAL_FONT  = Font(name="Arial", size=10)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)
# ─────────────────────────────────────────────────────────────────────────────


# ══════════════════════════════════════════════════════════════════════════════
#  Validation & helpers
# ══════════════════════════════════════════════════════════════════════════════

def validate_columns(df: pd.DataFrame) -> None:
    missing = REQUIRED_COLUMNS - set(df.columns)
    if missing:
        raise ValueError(
            f"Input file is missing required columns: {sorted(missing)}\n"
            f"Make sure you are loading output from citation_lookup.py"
        )


def safe_sheet_name(name: str) -> str:
    return re.sub(r'[\\/*?:\[\]]', '_', str(name))[:31]


def _is_valid_bucket(cat: str) -> bool:
    return cat in PUBLISHER_BUCKETS

def _valid_mask(df: pd.DataFrame) -> pd.Series:
    return df['Category'].apply(_is_valid_bucket)

def _invalid_mask(df: pd.DataFrame) -> pd.Series:
    return df['Category'] == 'Invalid'


# ══════════════════════════════════════════════════════════════════════════════
#  Classification
# ══════════════════════════════════════════════════════════════════════════════

def bucket_from_publisher(pub_str: str) -> str:
    """Map a publisher string to a PUBLISHER_BUCKETS label.
    Returns 'No Publisher' if the string is empty or missing."""
    raw = str(pub_str).strip()
    if not raw or raw.lower() in ('nan', 'none', ''):
        return 'No Publisher'
    pub = raw.lower()
    if any(k in pub for k in ('elsevier', 'cell press', 'pergamon')):
        return 'Elsevier'
    if 'wiley' in pub:
        return 'Wiley'
    if any(k in pub for k in ('american chemical society', 'acs')):
        return 'ACS'
    if any(k in pub for k in ('royal society of chemistry', 'rsc')):
        return 'RSC'
    if any(k in pub for k in ('nature', 'springer')):
        return 'Nature/Springer'
    if any(k in pub for k in ('aaas', 'advancement of science')):
        return 'Science/AAAS'
    return 'Other'


def classify(row: pd.Series) -> str:
    """
    Primary classification using status and title_match only.
      ❌ status or ❌ title_match → Invalid
      ✅ status                   → publisher bucket (from 'publisher' column)
                                    'No Publisher' if that column is empty
    """
    status  = str(row.get(COL_STATUS, ''))
    title_m = str(row.get(COL_TITLE_M, ''))

    if status == '❌' or title_m == '❌':
        return 'Invalid'

    return bucket_from_publisher(row.get(COL_PUBLISHER, ''))


def classify_journal_publisher(pub_str) -> str:
    """
    Classify the AI-extracted publisher from the 'journal_publisher' column
    into a PUBLISHER_BUCKETS label.
    Returns 'No Publisher' if the cell is empty or NaN.
    """
    if pub_str is None or (isinstance(pub_str, float) and pd.isna(pub_str)):
        return 'No Publisher'
    return bucket_from_publisher(pub_str)


def is_invalid_with_doi(row: pd.Series) -> bool:
    """
    Extended invalidity check: primary Invalid OR doi_match = ❌.
    Only counts DOI mismatch when a DOI was present in the original citation
    (doi_match = '–' means no DOI was present, so not counted).
    """
    if row.get('Category') == 'Invalid':
        return True
    if COL_DOI_M in row.index:
        return str(row.get(COL_DOI_M, '–')) == '❌'
    return False


# ══════════════════════════════════════════════════════════════════════════════
#  Styling helpers
# ══════════════════════════════════════════════════════════════════════════════

def _style_cell(cell, font=None, fill=None, alignment=None, border=None):
    if font:      cell.font      = font
    if fill:      cell.fill      = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border


def _autofit_columns(ws, min_width=10, max_width=55):
    _set_column_width(ws)


def _set_column_width(ws, width=15):
    for col_cells in ws.columns:
        ws.column_dimensions[
            get_column_letter(col_cells[0].column)
        ].width = width


def _write_dataframe(ws, df: pd.DataFrame, start_row: int,
                     start_col: int = 1) -> int:
    headers = [df.index.name or ""] + list(df.columns)
    for c, h in enumerate(headers, start_col):
        cell = ws.cell(row=start_row, column=c, value=h)
        _style_cell(cell, font=WHITE_FONT, fill=HEADER_FILL,
                    alignment=CENTER, border=THIN_BORDER)

    col_fill_map = {
        'Invalid':      INVALID_FILL,
        'No Publisher': NO_PUB_FILL,
    }
    for r_off, (idx, row_data) in enumerate(df.iterrows(), 1):
        excel_row = start_row + r_off
        idx_cell = ws.cell(row=excel_row, column=start_col, value=idx)
        _style_cell(idx_cell, font=BOLD_FONT, fill=SUBHEAD_FILL,
                    alignment=LEFT, border=THIN_BORDER)
        for c_off, (col_name, val) in enumerate(
                zip(df.columns, row_data), start_col + 1):
            cell = ws.cell(row=excel_row, column=c_off, value=val)
            fill = col_fill_map.get(col_name)
            _style_cell(cell, font=NORMAL_FONT, fill=fill,
                        alignment=CENTER, border=THIN_BORDER)

    return start_row + len(df) + 1


def _section_header(ws, row: int, text: str, n_cols: int):
    cell = ws.cell(row=row, column=1, value=text)
    _style_cell(cell,
                font=Font(name="Arial", bold=True, size=11, color="1F3864"),
                fill=SUBHEAD_FILL, alignment=LEFT)
    if n_cols > 1:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=n_cols)


# ══════════════════════════════════════════════════════════════════════════════
#  Sheet writers
# ══════════════════════════════════════════════════════════════════════════════

def _make_pivot(subset: pd.DataFrame) -> pd.DataFrame:
    counts = pd.pivot_table(
        subset, index=COL_AI_ENGINE, columns='Category',
        aggfunc='size', fill_value=0, observed=False
    )
    for col in TARGET_COLUMNS:
        if col not in counts.columns:
            counts[col] = 0
    counts = counts[[c for c in TARGET_COLUMNS if c in counts.columns]]
    counts.loc['TOTAL'] = counts.sum()
    counts.index.name   = 'AI Engine'
    return counts


def _write_percent_table(ws, counts_df: pd.DataFrame,
                          counts_start_row: int, pct_start_row: int,
                          start_col: int = 1) -> int:
    """
    Write a percentage table whose cells contain Excel formulas referencing
    the corresponding counts table above.

    Denominator for each data row  = SUM of that row across all columns
                                     (i.e. total citations for that engine).
    Denominator for the TOTAL row  = the TOTAL row's own SUM (same logic).

    Cells are formatted as Excel percentages (0.0%) so Excel renders them
    natively — no pre-computed strings.
    """
    from openpyxl.styles import numbers as xl_numbers

    PCT_FMT = '0.0%'
    col_fill_map = {
        'Invalid':      INVALID_FILL,
        'No Publisher': NO_PUB_FILL,
    }

    # ── Header row (mirrors counts header) ───────────────────────────────────
    headers = [counts_df.index.name or ""] + list(counts_df.columns)
    for c, h in enumerate(headers, start_col):
        cell = ws.cell(row=pct_start_row, column=c, value=h)
        _style_cell(cell, font=WHITE_FONT, fill=HEADER_FILL,
                    alignment=CENTER, border=THIN_BORDER)

    # ── Data rows ─────────────────────────────────────────────────────────────
    n_data_cols = len(counts_df.columns)
    first_data_col = start_col + 1                 # column after the index label
    last_data_col  = start_col + n_data_cols       # inclusive

    for r_off, idx in enumerate(counts_df.index, 1):
        pct_row    = pct_start_row + r_off
        counts_row = counts_start_row + r_off      # matching row in counts table

        # Index label cell
        idx_cell = ws.cell(row=pct_row, column=start_col, value=idx)
        _style_cell(idx_cell, font=BOLD_FONT, fill=SUBHEAD_FILL,
                    alignment=LEFT, border=THIN_BORDER)

        # Build column letters for the SUM denominator (all data cols, same counts row)
        first_letter = get_column_letter(first_data_col)
        last_letter  = get_column_letter(last_data_col)
        denom_ref    = f"SUM({first_letter}{counts_row}:{last_letter}{counts_row})"

        for c_off, col_name in enumerate(counts_df.columns, first_data_col):
            col_letter = get_column_letter(c_off)
            numerator  = f"{col_letter}{counts_row}"
            formula    = f"=IF({denom_ref}=0,\"\",{numerator}/{denom_ref})"

            cell = ws.cell(row=pct_row, column=c_off, value=formula)
            cell.number_format = PCT_FMT
            fill = col_fill_map.get(col_name)
            _style_cell(cell, font=NORMAL_FONT, fill=fill,
                        alignment=CENTER, border=THIN_BORDER)

    return pct_start_row + len(counts_df) + 1


def write_summary_sheet(writer, subset: pd.DataFrame,
                         sheet_name: str, title: str):
    counts = _make_pivot(subset)
    ws = writer.book.create_sheet(title=sheet_name)
    writer.sheets[sheet_name] = ws

    ws.cell(row=1, column=1, value=title)
    _style_cell(ws.cell(row=1, column=1),
                font=Font(name="Arial", bold=True, size=13, color="1F3864"))

    # ── Summary stats row (row 2) — formulas reference the counts table ───────
    # Counts table starts at row 5 (header) + 1 offset per engine row.
    # We'll write the stats as raw values here (they're just scalars, not per-row).
    total     = len(subset)
    valid_n   = _valid_mask(subset).sum()
    invalid_n = _invalid_mask(subset).sum()

    # Valid Rate and Invalid Rate written as Excel formulas once counts are placed.
    # We know the TOTAL row will be at counts_start_row + n_engines + 1.
    # Instead, keep scalars for Total/Valid/Invalid and use formulas for rates.
    stats = [
        ("Total",        total),
        ("✅ Valid",      valid_n),
        ("❌ Invalid",    invalid_n),
        ("Valid Rate",   f"={valid_n}/{total}" if total else "N/A"),
        ("Invalid Rate", f"={invalid_n}/{total}" if total else "N/A"),
    ]
    for c, (label, val) in enumerate(stats, 1):
        lc = ws.cell(row=2, column=c * 2 - 1, value=label)
        vc = ws.cell(row=2, column=c * 2,      value=val)
        if isinstance(val, str) and val.startswith("="):
            vc.number_format = '0.0%'
        _style_cell(lc, font=BOLD_FONT,   fill=SUBHEAD_FILL, border=THIN_BORDER)
        _style_cell(vc, font=NORMAL_FONT, fill=SUBHEAD_FILL,
                    alignment=CENTER, border=THIN_BORDER)

    _section_header(ws, 4, "Citation Counts by Publisher / Category",
                    len(counts.columns) + 1)
    counts_start_row = 5
    next_row = _write_dataframe(ws, counts, start_row=counts_start_row)

    _section_header(ws, next_row + 1, "Percentage Breakdown (%)",
                    len(counts.columns) + 1)
    pct_start_row = next_row + 2
    _write_percent_table(ws, counts, counts_start_row, pct_start_row)

    _autofit_columns(ws)


def write_invalidity_comparison(writer, df: pd.DataFrame):
    """
    Compares invalidity rates:
      Primary  — status + title_match only
      Extended — primary + citations with a mismatched DOI

    Broken down by AI engine and by topic.
    """
    has_doi_match = COL_DOI_M in df.columns

    sheet_name = "Invalidity_Comparison"
    ws = writer.book.create_sheet(title=sheet_name)
    writer.sheets[sheet_name] = ws

    ws.cell(row=1, column=1, value="Invalidity Rate Comparison")
    _style_cell(ws.cell(row=1, column=1),
                font=Font(name="Arial", bold=True, size=13, color="1F3864"))

    if not has_doi_match:
        ws.cell(row=2, column=1,
                value="doi_match column not found — extended analysis unavailable.")
        return

    df = df.copy()
    df['Invalid_Extended'] = df.apply(is_invalid_with_doi, axis=1)

    def _comparison_table(groupby_col: str, label: str,
                          start_row: int) -> int:
        _section_header(ws, start_row, label, 7)
        start_row += 1

        # Header — col positions: 1=group, 2=Total, 3=Prim, 4=%Prim, 5=Ext, 6=%Ext, 7=Addl
        headers = [groupby_col.replace('_', ' ').title(),
                   'Total', 'Invalid (Primary)', '% Primary',
                   'Invalid (+ DOI Mismatch)', '% Extended',
                   'Additional from DOI']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=c, value=h)
            _style_cell(cell, font=WHITE_FONT, fill=HEADER_FILL,
                        alignment=CENTER, border=THIN_BORDER)
        start_row += 1

        groups = sorted(df[groupby_col].dropna().unique())
        totals_row = {'total': 0, 'prim': 0, 'ext': 0}

        for grp_val in groups:
            grp   = df[df[groupby_col] == grp_val]
            total = len(grp)
            prim  = int(_invalid_mask(grp).sum())
            ext   = int(grp['Invalid_Extended'].sum())

            r = start_row  # current Excel row
            # Col B = Total, C = Primary count, E = Extended count
            B, C, E = f"B{r}", f"C{r}", f"E{r}"

            row_vals = [
                (1, str(grp_val), None,         LEFT),
                (2, total,        None,          CENTER),
                (3, prim,         INVALID_FILL,  CENTER),
                (4, f"=IF({B}=0,\"\",{C}/{B})", INVALID_FILL, CENTER),   # % Primary
                (5, ext,          WARN_FILL,     CENTER),
                (6, f"=IF({B}=0,\"\",{E}/{B})", WARN_FILL,   CENTER),   # % Extended
                (7, f"={E}-{C}",  None,          CENTER),                 # Additional
            ]
            for c, val, fill, align in row_vals:
                cell = ws.cell(row=r, column=c, value=val)
                if isinstance(val, str) and val.startswith('=') and '%' in headers[c-1]:
                    cell.number_format = '0.0%'
                _style_cell(cell, font=NORMAL_FONT, fill=fill,
                            alignment=align, border=THIN_BORDER)

            totals_row['total'] += total
            totals_row['prim']  += prim
            totals_row['ext']   += ext
            start_row += 1

        # Totals row — references its own B/C/E cells for the % formulas
        t, p, e = totals_row['total'], totals_row['prim'], totals_row['ext']
        r = start_row
        B, C, E = f"B{r}", f"C{r}", f"E{r}"
        totals_vals = [
            (1, 'TOTAL',                    None,         LEFT),
            (2, t,                           None,         CENTER),
            (3, p,                           INVALID_FILL, CENTER),
            (4, f"=IF({B}=0,\"\",{C}/{B})", INVALID_FILL, CENTER),
            (5, e,                           WARN_FILL,    CENTER),
            (6, f"=IF({B}=0,\"\",{E}/{B})", WARN_FILL,   CENTER),
            (7, f"={E}-{C}",                None,         CENTER),
        ]
        for c, val, fill, align in totals_vals:
            cell = ws.cell(row=r, column=c, value=val)
            if isinstance(val, str) and val.startswith('=') and '%' in headers[c-1]:
                cell.number_format = '0.0%'
            _style_cell(cell, font=BOLD_FONT, fill=SUBHEAD_FILL,
                        alignment=align, border=THIN_BORDER)
        return start_row + 2

    next_row = _comparison_table(COL_AI_ENGINE, "── By AI Engine ──", 3)
    _comparison_table(COL_TOPIC, "── By Topic ──", next_row)

    _autofit_columns(ws)


def write_year_analysis(writer, df: pd.DataFrame):
    sheet_name = "Publication_Years"
    ws = writer.book.create_sheet(title=sheet_name)
    writer.sheets[sheet_name] = ws

    next_row = 1
    for label, subset in [("All", df), ("Valid", df[_valid_mask(df)])]:
        if subset.empty:
            ws.cell(row=next_row, column=1,
                    value=f"No {label.lower()} citations found.")
            next_row += 2
            continue

        pivot = pd.pivot_table(
            subset, index=COL_YEAR, columns=COL_AI_ENGINE,
            aggfunc='size', fill_value=0, observed=False
        )
        pivot.index = pivot.index.astype(str)
        pivot = pivot.sort_index(ascending=False)
        pivot['Grand Total'] = pivot.sum(axis=1)
        pivot.index.name = 'Year'

        ws.cell(row=next_row, column=1,
                value=f"Publication Years — {label} Citations")
        _style_cell(ws.cell(row=next_row, column=1),
                    font=Font(name="Arial", bold=True, size=13, color="1F3864"))
        next_row += 1
        next_row = _write_dataframe(ws, pivot, start_row=next_row)
        next_row += 1

    _autofit_columns(ws)


def write_top_journals(writer, df: pd.DataFrame, top_n: int = 5):
    UNKNOWN = {'Unknown Journal', 'Unknown', '', 'No DOI'}
    PREPRINTS = {'ChemRxiv', 'Research Square', 'arXiv',
                 'bioRxiv', 'medRxiv', 'SSRN'}

    def _clean(subset):
        return subset[subset[COL_JOURNAL].notna() &
                      ~subset[COL_JOURNAL].isin(UNKNOWN)]

    for label, subset in [
        ("All",   _clean(df)),
        ("Valid", _clean(df[_valid_mask(df) & ~df[COL_JOURNAL].isin(PREPRINTS)]))
    ]:
        sheet_name = f"Top_{top_n}_Journals_{label}"
        ws = writer.book.create_sheet(title=sheet_name)
        writer.sheets[sheet_name] = ws
        ws.cell(row=1, column=1, value=f"Top {top_n} Journals — {label} Citations")
        _style_cell(ws.cell(row=1, column=1),
                    font=Font(name="Arial", bold=True, size=13, color="1F3864"))

        if subset.empty:
            ws.cell(row=2, column=1, value="No citations found.")
            continue

        start_row = 3
        for eng_label, eng_df in [("ALL ENGINES COMBINED", subset)] + [
            (eng, subset[subset[COL_AI_ENGINE] == eng])
            for eng in sorted(df[COL_AI_ENGINE].dropna().unique())
        ]:
            _section_header(ws, start_row, f"── {eng_label} ──", 2)
            start_row += 1
            for c, h in enumerate(['Journal', 'Count'], 1):
                cell = ws.cell(row=start_row, column=c, value=h)
                _style_cell(cell, font=WHITE_FONT, fill=HEADER_FILL,
                            alignment=CENTER, border=THIN_BORDER)
            start_row += 1
            top = (eng_df[COL_JOURNAL].value_counts()
                   .head(top_n).reset_index())
            top.columns = ['Journal', 'Count']
            for _, jrow in top.iterrows():
                for c, val in enumerate([jrow['Journal'], jrow['Count']], 1):
                    cell = ws.cell(row=start_row, column=c, value=val)
                    _style_cell(cell, font=NORMAL_FONT,
                                alignment=LEFT if c == 1 else CENTER,
                                border=THIN_BORDER)
                start_row += 1
            start_row += 2
        _autofit_columns(ws)


# ══════════════════════════════════════════════════════════════════════════════
#  Publisher preference sheets
# ══════════════════════════════════════════════════════════════════════════════

def _build_paired_pivot(subset: pd.DataFrame,
                        groupby_col: str) -> pd.DataFrame:
    """
    Build a counts table with paired columns for every publisher bucket:
      <Publisher> (Val)  — from validated publisher ('publisher' column → Category)
      <Publisher> (Orig) — from AI-extracted publisher ('journal_publisher' column → OrigCategory)

    The 'Invalid' bucket is appended as a single column at the end
    (validated only — invalid citations have no meaningful 'original' publisher).

    Returns a DataFrame with a TOTAL row appended.
    """
    val_counts  = pd.pivot_table(
        subset, index=groupby_col, columns='Category',
        aggfunc='size', fill_value=0, observed=False
    )
    orig_counts = pd.pivot_table(
        subset, index=groupby_col, columns='OrigCategory',
        aggfunc='size', fill_value=0, observed=False
    )

    # Ensure all publisher buckets present in both
    for col in PUBLISHER_BUCKETS:
        if col not in val_counts.columns:
            val_counts[col]  = 0
        if col not in orig_counts.columns:
            orig_counts[col] = 0

    # Build interleaved paired columns for publisher buckets
    paired_cols = {}
    for pub in PUBLISHER_BUCKETS:
        paired_cols[f"{pub} (Val)"]  = val_counts.get(pub,  0)
        paired_cols[f"{pub} (Orig)"] = orig_counts.get(pub, 0)

    # Append non-publisher buckets (validated only)
    for bucket in INVALID_BUCKETS:
        paired_cols[bucket] = val_counts.get(bucket, 0)

    result = pd.DataFrame(paired_cols, index=val_counts.index)
    result.index.name = groupby_col
    result.loc['TOTAL'] = result.sum()
    return result


def write_publisher_preference(writer, df: pd.DataFrame):
    """
    One sheets showing publisher/category distribution:
      Publisher_Preference_All        — full dataset
      

    Sheet shows counts and percentages broken down by:
      - AI engine (rows) × publisher bucket (columns)
      - Topic (rows)     × publisher bucket (columns)

    For each publisher bucket, two sub-columns are shown:
      (Val)  — publisher from 'publisher' column (validated via Crossref/lookup)
      (Orig) — publisher from 'journal_publisher' column (AI-extracted)
    """
    # Colour band: alternate light blue / light green per publisher pair
    PAIR_FILLS = [
        PatternFill("solid", fgColor="DCE6F1"),  # light blue
        PatternFill("solid", fgColor="EBF1DE"),  # light green
    ]
    INVALID_COL_FILL = PatternFill("solid", fgColor="F2F2F2")  # light grey for Invalid column

    def _write_paired_dataframe(ws, df_data: pd.DataFrame,
                                start_row: int, start_col: int = 1) -> int:
        """
        Write a paired-column DataFrame with:
          - A super-header row grouping Val/Orig pairs under the publisher name
          - A sub-header row with (Val) / (Orig) labels
          - Data rows with alternating fill per publisher pair
        """
        cols        = list(df_data.columns)
        index_vals  = list(df_data.index)
        n_pub_pairs = len(PUBLISHER_BUCKETS)        # paired columns
        n_invalid   = len(INVALID_BUCKETS)          # single columns at end

        # ── Row 1: super-headers ──────────────────────────────────────────────
        # Index label cell
        idx_cell = ws.cell(row=start_row, column=start_col,
                           value=df_data.index.name or "")
        ws.merge_cells(start_row=start_row, start_column=start_col,
                       end_row=start_row + 1, end_column=start_col)
        _style_cell(idx_cell, font=WHITE_FONT, fill=HEADER_FILL,
                    alignment=CENTER, border=THIN_BORDER)

        col_cursor = start_col + 1
        for i, pub in enumerate(PUBLISHER_BUCKETS):
            pair_fill = PatternFill("solid", fgColor=(
                "2E75B6" if i % 2 == 0 else "375623"
            ))
            pair_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            cell = ws.cell(row=start_row, column=col_cursor, value=pub)
            ws.merge_cells(start_row=start_row, start_column=col_cursor,
                           end_row=start_row,   end_column=col_cursor + 1)
            _style_cell(cell, font=pair_font, fill=pair_fill,
                        alignment=CENTER, border=THIN_BORDER)
            col_cursor += 2

        for bucket in INVALID_BUCKETS:
            cell = ws.cell(row=start_row, column=col_cursor, value=bucket)
            ws.merge_cells(start_row=start_row, start_column=col_cursor,
                           end_row=start_row + 1, end_column=col_cursor)
            _style_cell(cell, font=BOLD_FONT, fill=INVALID_FILL,
                        alignment=CENTER, border=THIN_BORDER)
            col_cursor += 1

        # ── Row 2: sub-headers (Val) / (Orig) ────────────────────────────────
        sub_row    = start_row + 1
        col_cursor = start_col + 1
        for i in range(n_pub_pairs):
            pair_fill = PAIR_FILLS[i % 2]
            for sub_label in ("Val", "Orig"):
                cell = ws.cell(row=sub_row, column=col_cursor, value=sub_label)
                _style_cell(cell, font=BOLD_FONT, fill=pair_fill,
                            alignment=CENTER, border=THIN_BORDER)
                col_cursor += 1
        # invalid bucket sub-headers already merged above; skip their columns
        for _ in INVALID_BUCKETS:
            col_cursor += 1   # borders/fill handled by merge above

        # ── Data rows ─────────────────────────────────────────────────────────
        data_start = start_row + 2
        for r_off, idx_val in enumerate(index_vals):
            excel_row  = data_start + r_off
            row_series = df_data.loc[idx_val]

            # Index cell
            idx_c = ws.cell(row=excel_row, column=start_col, value=idx_val)
            _style_cell(idx_c, font=BOLD_FONT, fill=SUBHEAD_FILL,
                        alignment=LEFT, border=THIN_BORDER)

            col_cursor = start_col + 1
            for i, pub in enumerate(PUBLISHER_BUCKETS):
                pair_fill = PAIR_FILLS[i % 2]
                for suffix in ("Val", "Orig"):
                    col_name = f"{pub} ({suffix})"
                    val  = row_series.get(col_name, 0)
                    cell = ws.cell(row=excel_row, column=col_cursor, value=val)
                    _style_cell(cell, font=NORMAL_FONT, fill=pair_fill,
                                alignment=CENTER, border=THIN_BORDER)
                    col_cursor += 1

            for bucket in INVALID_BUCKETS:
                val  = row_series.get(bucket, 0)
                cell = ws.cell(row=excel_row, column=col_cursor, value=val)
                _style_cell(cell, font=NORMAL_FONT, fill=INVALID_COL_FILL,
                            alignment=CENTER, border=THIN_BORDER)
                col_cursor += 1

        return data_start + len(index_vals)

    def _write_paired_percent_table(ws, counts_df: pd.DataFrame,
                                    counts_start_row: int, pct_start_row: int,
                                    start_col: int = 1) -> int:
        """
        Write a percentage table matching the layout of _write_paired_dataframe,
        but every data cell contains an Excel formula referencing the corresponding
        cell in the counts table above.

        Denominator = SUM of all (Val) columns + Invalid columns in that counts row
        (i.e. the total number of citations for that engine/topic).
        Cells are formatted as Excel percentages (0.0%).
        """
        PCT_FMT     = '0.0%'
        index_vals  = list(counts_df.index)
        n_pub_pairs = len(PUBLISHER_BUCKETS)
        n_invalid   = len(INVALID_BUCKETS)

        # Build the column numbers of the Val columns + Invalid columns
        # (these form the denominator — total citations per row)
        # Layout: start_col=1 (index), then pairs of (Val, Orig) per publisher, then Invalids
        val_col_numbers   = [start_col + 1 + i * 2 for i in range(n_pub_pairs)]   # every Val col
        invalid_col_start = start_col + 1 + n_pub_pairs * 2
        invalid_col_numbers = [invalid_col_start + j for j in range(n_invalid)]
        denom_col_numbers = val_col_numbers + invalid_col_numbers

        # ── Super-header row (mirrors counts super-header) ────────────────────
        idx_cell = ws.cell(row=pct_start_row, column=start_col,
                           value=counts_df.index.name or "")
        ws.merge_cells(start_row=pct_start_row, start_column=start_col,
                       end_row=pct_start_row + 1, end_column=start_col)
        _style_cell(idx_cell, font=WHITE_FONT, fill=HEADER_FILL,
                    alignment=CENTER, border=THIN_BORDER)

        col_cursor = start_col + 1
        for i, pub in enumerate(PUBLISHER_BUCKETS):
            pair_fill = PatternFill("solid", fgColor=(
                "2E75B6" if i % 2 == 0 else "375623"
            ))
            pair_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            cell = ws.cell(row=pct_start_row, column=col_cursor, value=pub)
            ws.merge_cells(start_row=pct_start_row, start_column=col_cursor,
                           end_row=pct_start_row,   end_column=col_cursor + 1)
            _style_cell(cell, font=pair_font, fill=pair_fill,
                        alignment=CENTER, border=THIN_BORDER)
            col_cursor += 2

        for bucket in INVALID_BUCKETS:
            cell = ws.cell(row=pct_start_row, column=col_cursor, value=bucket)
            ws.merge_cells(start_row=pct_start_row, start_column=col_cursor,
                           end_row=pct_start_row + 1, end_column=col_cursor)
            _style_cell(cell, font=BOLD_FONT, fill=INVALID_FILL,
                        alignment=CENTER, border=THIN_BORDER)
            col_cursor += 1

        # ── Sub-header row ────────────────────────────────────────────────────
        sub_row    = pct_start_row + 1
        col_cursor = start_col + 1
        for i in range(n_pub_pairs):
            pair_fill = PAIR_FILLS[i % 2]
            for sub_label in ("Val", "Orig"):
                cell = ws.cell(row=sub_row, column=col_cursor, value=sub_label)
                _style_cell(cell, font=BOLD_FONT, fill=pair_fill,
                            alignment=CENTER, border=THIN_BORDER)
                col_cursor += 1
        for _ in INVALID_BUCKETS:
            col_cursor += 1

        # ── Data rows ─────────────────────────────────────────────────────────
        data_start = pct_start_row + 2
        for r_off, idx_val in enumerate(index_vals):
            pct_row    = data_start + r_off
            counts_row = counts_start_row + 2 + r_off   # +2 for the two header rows

            # Build denominator formula: sum of Val cols + Invalid cols in counts table
            denom_refs = '+'.join(
                f"{get_column_letter(cn)}{counts_row}" for cn in denom_col_numbers
            )
            denom_expr = f"({denom_refs})"

            # Index label
            idx_c = ws.cell(row=pct_row, column=start_col, value=idx_val)
            _style_cell(idx_c, font=BOLD_FONT, fill=SUBHEAD_FILL,
                        alignment=LEFT, border=THIN_BORDER)

            col_cursor = start_col + 1
            for i, pub in enumerate(PUBLISHER_BUCKETS):
                pair_fill = PAIR_FILLS[i % 2]
                for suffix in ("Val", "Orig"):
                    counts_col_letter = get_column_letter(col_cursor)
                    numerator = f"{counts_col_letter}{counts_row}"
                    formula   = f"=IF({denom_expr}=0,\"\",{numerator}/{denom_expr})"
                    cell = ws.cell(row=pct_row, column=col_cursor, value=formula)
                    cell.number_format = PCT_FMT
                    _style_cell(cell, font=NORMAL_FONT, fill=pair_fill,
                                alignment=CENTER, border=THIN_BORDER)
                    col_cursor += 1

            for _ in INVALID_BUCKETS:
                counts_col_letter = get_column_letter(col_cursor)
                numerator = f"{counts_col_letter}{counts_row}"
                formula   = f"=IF({denom_expr}=0,\"\",{numerator}/{denom_expr})"
                cell = ws.cell(row=pct_row, column=col_cursor, value=formula)
                cell.number_format = PCT_FMT
                _style_cell(cell, font=NORMAL_FONT, fill=INVALID_COL_FILL,
                            alignment=CENTER, border=THIN_BORDER)
                col_cursor += 1

        return data_start + len(index_vals)

    # ── Sheet loop ─────────────────────────────────────────────────────────────
    for label, subset in [
        ("All",        df)
    ]:
        sheet_name = f"Publisher_Preference_{label}"
        ws = writer.book.create_sheet(title=sheet_name)
        writer.sheets[sheet_name] = ws

        title = (
            "Publisher Preference — All Citations" if label == "All"
            else "Publisher Preference — Valid Citations Only"
        )
        ws.cell(row=1, column=1, value=title)
        _style_cell(ws.cell(row=1, column=1),
                    font=Font(name="Arial", bold=True, size=13, color="1F3864"))

        if subset.empty:
            ws.cell(row=2, column=1, value="No data available.")
            continue

        # Legend
        legend_row = 2
        ws.cell(row=legend_row, column=1,
                value="(Val) = publisher from 'publisher' column (Crossref/lookup validated)"
                      "     (Orig) = publisher from 'journal_publisher' column (AI-extracted)")
        _style_cell(ws.cell(row=legend_row, column=1),
                    font=Font(name="Arial", italic=True, size=9, color="595959"))

        next_row = 4

        for groupby_col, group_label in [
            (COL_AI_ENGINE, "By AI Engine"),
            (COL_TOPIC,     "By Topic"),
        ]:
            n_display_cols = len(PUBLISHER_BUCKETS) * 2 + len(INVALID_BUCKETS) + 1
            _section_header(ws, next_row, f"── {group_label} ──", n_display_cols)
            next_row += 1

            counts = _build_paired_pivot(subset, groupby_col)
            counts.index.name = group_label

            _section_header(ws, next_row, "Counts", n_display_cols)
            next_row += 1
            counts_start_row = next_row
            next_row = _write_paired_dataframe(ws, counts, start_row=counts_start_row)

            next_row += 1
            _section_header(ws, next_row, "Percentages (%)", n_display_cols)
            next_row += 1
            next_row = _write_paired_percent_table(
                ws, counts,
                counts_start_row=counts_start_row,
                pct_start_row=next_row
            )
            next_row += 2

        _autofit_columns(ws)


def write_combined_topic_sheet(writer, df_all: pd.DataFrame,
                               df_ni: pd.DataFrame,
                               sheet_name: str, topic: str):
    """
    Write one sheet per topic combining the Full Dataset summary (top)
    and Valid Only summary (below) — same structure as write_summary_sheet,
    stacked vertically with a blank row between them.
    """
    ws = writer.book.create_sheet(title=sheet_name)
    writer.sheets[sheet_name] = ws

    def _write_block(subset: pd.DataFrame, title: str, start_row: int) -> int:
        counts = _make_pivot(subset)

        ws.cell(row=start_row, column=1, value=title)
        _style_cell(ws.cell(row=start_row, column=1),
                    font=Font(name="Arial", bold=True, size=13, color="1F3864"))

        total     = len(subset)
        valid_n   = _valid_mask(subset).sum()
        invalid_n = _invalid_mask(subset).sum()
        stats = [
            ("Total",        total),
            ("✅ Valid",      valid_n),
            ("❌ Invalid",    invalid_n),
            ("Valid Rate",   f"={valid_n}/{total}" if total else "N/A"),
            ("Invalid Rate", f"={invalid_n}/{total}" if total else "N/A"),
        ]
        for c, (label, val) in enumerate(stats, 1):
            lc = ws.cell(row=start_row + 1, column=c * 2 - 1, value=label)
            vc = ws.cell(row=start_row + 1, column=c * 2,      value=val)
            if isinstance(val, str) and val.startswith("="):
                vc.number_format = '0.0%'
            _style_cell(lc, font=BOLD_FONT,   fill=SUBHEAD_FILL, border=THIN_BORDER)
            _style_cell(vc, font=NORMAL_FONT, fill=SUBHEAD_FILL,
                        alignment=CENTER, border=THIN_BORDER)

        _section_header(ws, start_row + 3,
                        "Citation Counts by Publisher / Category",
                        len(counts.columns) + 1)
        counts_start_row = start_row + 4
        next_row = _write_dataframe(ws, counts, start_row=counts_start_row)

        _section_header(ws, next_row + 1, "Percentage Breakdown (%)",
                        len(counts.columns) + 1)
        pct_start_row = next_row + 2
        _write_percent_table(ws, counts, counts_start_row, pct_start_row)
        return pct_start_row + len(counts) + 2

    if not df_all.empty:
        end_row = _write_block(df_all, f"Topic: {topic} — Full Dataset", 1)
    else:
        ws.cell(row=1, column=1, value=f"Topic: {topic} — Full Dataset")
        _style_cell(ws.cell(row=1, column=1),
                    font=Font(name="Arial", bold=True, size=13, color="1F3864"))
        ws.cell(row=2, column=1, value="No data available.")
        end_row = 4

    if not df_ni.empty:
        _write_block(df_ni, f"Topic: {topic} — Valid Only", end_row + 1)
    else:
        ws.cell(row=end_row + 1, column=1,
                value=f"Topic: {topic} — Valid Only")
        _style_cell(ws.cell(row=end_row + 1, column=1),
                    font=Font(name="Arial", bold=True, size=13, color="1F3864"))
        ws.cell(row=end_row + 2, column=1, value="No valid data available.")

    _autofit_columns(ws)


# ══════════════════════════════════════════════════════════════════════════════
#  Main
# ══════════════════════════════════════════════════════════════════════════════

def _load_and_classify(file_path: str) -> pd.DataFrame | None:
    """Load one Excel file, validate, clean, classify. Returns DataFrame."""
    try:
        df = pd.read_excel(file_path)
    except Exception as e:
        print(f"  ❌ Could not read {os.path.basename(file_path)}: {e}")
        return None

    try:
        validate_columns(df)
    except ValueError as e:
        print(f"  ❌ {e}")
        return None

    for col in (COL_JOURNAL, COL_PUBLISHER, COL_FOUND, COL_ORIG):
        if col in df.columns:
            df[col] = df[col].fillna('').astype(str).apply(html.unescape)

    df['Category'] = df.apply(classify, axis=1)
    # AI-extracted publisher bucket — from the dedicated 'journal_publisher' column
    if COL_JOURNAL_PUB in df.columns:
        df['OrigCategory'] = df[COL_JOURNAL_PUB].apply(classify_journal_publisher)
    else:
        df['OrigCategory'] = 'Other'
    # Tag which source file this row came from (filename without extension)
    df['_source'] = os.path.splitext(os.path.basename(file_path))[0]
    return df


def main():
    root = tk.Tk()
    root.withdraw()

    print("Select one or more Citation Lookup Results Excel files…")
    file_paths = filedialog.askopenfilenames(
        title="Select Lookup Results Excel (one or multiple)",
        filetypes=[("Excel Files", "*.xlsx")]
    )
    if not file_paths:
        print("No files selected. Exiting.")
        return

    frames = []
    for fp in file_paths:
        print(f"  Loading: {os.path.basename(fp)}")
        df_single = _load_and_classify(fp)
        if df_single is not None:
            frames.append(df_single)
            print(f"    → {len(df_single)} rows, "
                  f"{df_single[COL_AI_ENGINE].nunique()} engine(s), "
                  f"{df_single[COL_TOPIC].nunique()} topic(s)")

    if not frames:
        print("No valid data loaded. Exiting.")
        return

    df = pd.concat(frames, ignore_index=True)
    print(f"\n  Combined: {len(df)} total rows from {len(frames)} file(s)")

    # Valid (non-invalid) subset
    df_ni = df[df['Category'] != 'Invalid'].copy()

    # Output filename — single file keeps its name, multiple gets "Combined"
    if len(frames) == 1:
        out_stem = os.path.splitext(os.path.basename(file_paths[0]))[0]
        out_dir  = os.path.dirname(file_paths[0])
    else:
        out_stem = "Combined"
        out_dir  = os.path.dirname(file_paths[0])
    output_path = os.path.join(out_dir, f"{out_stem}_Analysis.xlsx")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        if 'Sheet' in writer.book.sheetnames:
            del writer.book['Sheet']

        # 1. Invalidity comparison
        write_invalidity_comparison(writer, df)

        # 2. Publisher preference — All
        write_publisher_preference(writer, df)

        # 3. Publication Years (All + Valid combined into one sheet)
        write_year_analysis(writer, df)

        # 4 & 5. Top journals (All, then Valid)
        write_top_journals(writer, df, top_n=5)

        # 6+. Combined topic sheets (All + Valid side-by-side, one sheet per topic)
        for topic in sorted(df[COL_TOPIC].dropna().unique()):
            sheet_name = safe_sheet_name(str(topic))
            df_topic_all = df[df[COL_TOPIC] == topic]
            df_topic_ni  = df_ni[df_ni[COL_TOPIC] == topic] if not df_ni.empty else pd.DataFrame()
            write_combined_topic_sheet(
                writer, df_topic_all, df_topic_ni,
                sheet_name, str(topic)
            )





    # Console summary
    total    = len(df)
    valid_n  = _valid_mask(df).sum()
    invalid_n = _invalid_mask(df).sum()

    print(f"\n{'━'*65}")
    print(f"  Files loaded        : {len(frames)}")
    print(f"  Total citations     : {total}")
    print(f"  ✅ Valid            : {valid_n}  ({100*valid_n/total:.1f}%)")
    print(f"  ❌ Invalid (primary): {invalid_n}  ({100*invalid_n/total:.1f}%)")
    if COL_DOI_M in df.columns:
        ext_n = df.apply(is_invalid_with_doi, axis=1).sum()
        print(f"  ❌ Invalid (+DOI)  : {ext_n}  ({100*ext_n/total:.1f}%)"
              f"  (+{ext_n - invalid_n} from DOI mismatch)")
    print(f"  Valid subset        : {len(df_ni)}/{total}")
    print(f"{'━'*65}")
    print(f"  Output → {output_path}")
    print(f"{'━'*65}\n")


if __name__ == "__main__":
    main()