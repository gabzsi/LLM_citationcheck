import os
import re
import time
import requests
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIGURATION ─────────────────────────────────────────────────────────────
USER_EMAIL = "your.name@email.com" # Required by Unpaywall

COL_AI_ENGINE = 'ai_engine'
COL_DOI       = 'doi'
COL_STATUS    = 'status'
COL_IS_OA     = 'Open_Access'

# Styling
HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
SUBHEAD_FILL = PatternFill("solid", fgColor="D9E1F2")
OA_FILL      = PatternFill("solid", fgColor="C6EFCE")
CLOSED_FILL  = PatternFill("solid", fgColor="FFC7CE")
WHITE_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BOLD_FONT    = Font(name="Arial", bold=True, size=10)
NORMAL_FONT  = Font(name="Arial", size=10)
CENTER       = Alignment(horizontal="center", vertical="center")
LEFT         = Alignment(horizontal="left",   vertical="center")
THIN_BORDER  = Border(left=Side(style="thin"), right=Side(style="thin"), 
                     top=Side(style="thin"), bottom=Side(style="thin"))

# ── Logic ─────────────────────────────────────────────────────────────────────

def clean_doi(doi_val):
    if not doi_val or pd.isna(doi_val): return None
    match = re.search(r'(10\.\d{4,9}/[-._;()/:A-Z0-9]+)', str(doi_val), re.IGNORECASE)
    return match.group(1) if match else None

def get_oa_data(doi):
    target = clean_doi(doi)
    if not target: return "NO", "closed"
    try:
        r = requests.get(f"https://api.unpaywall.org/v2/{target}", 
                         params={'email': USER_EMAIL}, timeout=10)
        if r.status_code == 200:
            data = r.json()
            return ("YES" if data.get('is_oa') else "NO"), data.get('oa_status', 'closed')
        return "NO", "closed"
    except:
        return "NO", "closed"

def _style_cell(cell, font=None, fill=None, alignment=None, border=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border

# ── Analysis Writer ───────────────────────────────────────────────────────────

def write_oa_comparison_table(ws, df, start_row, title):
    """Creates the summary table: Engine vs OA Status."""
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12)
    start_row += 1
    
    # Create Pivot (AI Engine as rows, YES/NO as columns)
    pivot = pd.pivot_table(df, index=COL_AI_ENGINE, columns=COL_IS_OA, 
                           aggfunc='size', fill_value=0)
    
    # Safely ensure both columns exist to avoid indexing errors
    for col in ["YES", "NO"]:
        if col not in pivot.columns: pivot[col] = 0
    
    pivot = pivot.reindex(columns=["YES", "NO"])
    pivot.columns = ['Open Access', 'Paid/Closed']
    pivot.loc['TOTAL'] = pivot.sum()

    headers = ["AI Engine", "Open Access Citations", "Paid Citations", "% Open Access"]
    for c, h in enumerate(headers, 1):
        _style_cell(ws.cell(row=start_row, column=c, value=h), 
                    font=WHITE_FONT, fill=HEADER_FILL, border=THIN_BORDER, alignment=CENTER)

    curr_row = start_row
    for r_idx, (engine, row_data) in enumerate(pivot.iterrows(), 1):
        curr_row = start_row + r_idx
        oa_count = row_data['Open Access']
        paid_count = row_data['Paid/Closed']
        oa_ref   = f"B{curr_row}"
        paid_ref = f"C{curr_row}"
        pct = f"=IF({oa_ref}+{paid_ref}=0,\"\",{oa_ref}/({oa_ref}+{paid_ref}))"

        vals = [engine, oa_count, paid_count, pct]
        for c_idx, v in enumerate(vals, 1):
            cell = ws.cell(row=curr_row, column=c_idx, value=v)
            if c_idx == 4:
                cell.number_format = '0.0%'
            is_total = (engine == 'TOTAL')
            _style_cell(cell, font=BOLD_FONT if is_total else NORMAL_FONT, 
                        fill=SUBHEAD_FILL if is_total else None, border=THIN_BORDER,
                        alignment=CENTER if c_idx > 1 else LEFT)
    return curr_row + 2

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    root = tk.Tk(); root.withdraw()
    print("Select your Citation Results Excel files…")
    file_paths = filedialog.askopenfilenames(filetypes=[("Excel Files", "*.xlsx")])
    if not file_paths: return

    processed_frames = []
    for fp in file_paths:
        print(f"Reading: {os.path.basename(fp)}")
        df = pd.read_excel(fp)
        
        # Filter: Ignore citations marked as hallucinations (❌)
        valid_df = df[df[COL_STATUS] != '❌'].copy()
        if valid_df.empty: continue

        print(f"  Checking Unpaywall API for {len(valid_df)} citations...")
        results = []
        for i, doi in enumerate(valid_df[COL_DOI]):
            is_oa, oa_type = get_oa_data(doi)
            results.append({COL_IS_OA: is_oa, 'OA_Type': oa_type})
            if (i+1) % 10 == 0: print(f"    Done {i+1}...")
            time.sleep(0.15)
        
        res_df = pd.DataFrame(results)
        valid_df[COL_IS_OA] = res_df[COL_IS_OA].values
        valid_df['OA_Type'] = res_df['OA_Type'].values
        processed_frames.append(valid_df)

    if not processed_frames: return
    final_df = pd.concat(processed_frames, ignore_index=True)
    out_path = os.path.join(os.path.dirname(file_paths[0]), "AI_Engine_OA_Comparison_0317.xlsx")

    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        # Sheet 1: Summary Table
        ws_sum = writer.book.create_sheet("Summary_Analysis")
        write_oa_comparison_table(ws_sum, final_df, 1, "Citations: Open Access vs Paid per AI Engine")
        
        # Autofit summary
        for col in ws_sum.columns:
            ws_sum.column_dimensions[get_column_letter(col[0].column)].width = 25

        # Sheet 2: Detailed Log
        final_df.to_excel(writer, sheet_name="Detailed_Data", index=False)
        ws_det = writer.sheets["Detailed_Data"]
        oa_col_idx = final_df.columns.get_loc(COL_IS_OA) + 1
        for row in range(2, ws_det.max_row + 1):
            cell = ws_det.cell(row=row, column=oa_col_idx)
            cell.fill = OA_FILL if cell.value == "YES" else CLOSED_FILL

    print(f"\nDone! Analysis saved to: {out_path}")

if __name__ == "__main__":
    main()