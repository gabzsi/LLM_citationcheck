"""
━━━━━━━━━━━━━━━━━━━━━━━
Finds duplicate DOIs within and across runs, with two extended analyses.

Run detection:
  - Files with '0317' in the filename → Run 1
  - Files with '0319' in the filename → Run 2
  - If ambiguous, a dialog will ask you to assign manually.

Output Excel sheets:
  1.  Summary                — counts at a glance for all analyses
  2.  Run1_Duplicates        — DOIs appearing >1× anywhere within Run 1
  3.  Run2_Duplicates        — DOIs appearing >1× anywhere within Run 2
  4.  Cross_Run_Duplicates   — DOIs present in BOTH runs (any engine)
  5.  MultiEngine_Run1       — DOIs cited by >1 distinct engine on Run 1 day
  6.  MultiEngine_Run2       — DOIs cited by >1 distinct engine on Run 2 day
  7.  SameEngine_CrossRun    — same engine cited same DOI in both Run 1 & Run 2

Columns kept throughout: ai_engine, topic, doi, original_citation
"""

import os
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Column names ──────────────────────────────────────────────────────────────
COL_AI_ENGINE = 'ai_engine'
COL_TOPIC     = 'topic'
COL_DOI       = 'doi'
COL_ORIG      = 'original_citation'
OUTPUT_COLS   = [COL_AI_ENGINE, COL_TOPIC, COL_DOI, COL_ORIG]

# ── Palette ───────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
RUN1_FILL   = PatternFill("solid", fgColor="D9E1F2")
RUN2_FILL   = PatternFill("solid", fgColor="E2EFDA")
CROSS_FILL  = PatternFill("solid", fgColor="FCE4D6")
MULTI_FILL  = PatternFill("solid", fgColor="EAD1DC")
SENG_FILL   = PatternFill("solid", fgColor="FFF2CC")
ALT_FILL    = PatternFill("solid", fgColor="F2F2F2")
GRP_FILL    = PatternFill("solid", fgColor="404040")

ENGINE_FILLS = [
    PatternFill("solid", fgColor="BDD7EE"),
    PatternFill("solid", fgColor="C6EFCE"),
    PatternFill("solid", fgColor="FFDAB9"),
    PatternFill("solid", fgColor="E6CCFF"),
    PatternFill("solid", fgColor="FFFACD"),
    PatternFill("solid", fgColor="FFD1DC"),
]

# ── Fonts / alignment ─────────────────────────────────────────────────────────
WHITE_FONT = Font(name="Arial", bold=True,   color="FFFFFF", size=10)
BOLD_FONT  = Font(name="Arial", bold=True,   size=10)
NORM_FONT  = Font(name="Arial", size=10)
ITAL_FONT  = Font(name="Arial", italic=True, size=10, color="595959")
GRP_FONT   = Font(name="Arial", bold=True,   color="FFFFFF", size=9)
ENG_FONT   = Font(name="Arial", bold=True,   size=11, color="1F3864")
CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN       = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"),  bottom=Side(style="thin"))


# ══════════════════════════════════════════════════════════════════════════════
#  Load & normalise
# ══════════════════════════════════════════════════════════════════════════════

def _detect_run(filepath):
    name = os.path.basename(filepath)
    h17, h19 = '0317' in name, '0319' in name
    if h17 and not h19: return '1'
    if h19 and not h17: return '2'
    return None


def _load_file(filepath):
    try:
        df = pd.read_excel(filepath)
    except Exception as e:
        print(f"  Could not read {os.path.basename(filepath)}: {e}")
        return None
    for c in OUTPUT_COLS:
        if c not in df.columns:
            print(f"  Column '{c}' missing in {os.path.basename(filepath)} — filled blank")
            df[c] = ''
    df[COL_DOI]       = df[COL_DOI].fillna('').astype(str).str.strip().str.lower()
    df[COL_AI_ENGINE] = df[COL_AI_ENGINE].fillna('').astype(str).str.strip()
    return df[OUTPUT_COLS].copy()


def _valid(df):
    return df[(df[COL_DOI] != '') & (df[COL_DOI] != 'nan')].copy()


# ══════════════════════════════════════════════════════════════════════════════
#  Analysis functions
# ══════════════════════════════════════════════════════════════════════════════

def _within_run_dups(df):
    v = _valid(df)
    mask = v.duplicated(subset=[COL_DOI], keep=False)
    return v[mask].sort_values(COL_DOI).reset_index(drop=True)


def _cross_run_dups(df1, df2):
    shared = set(_valid(df1)[COL_DOI]) & set(_valid(df2)[COL_DOI])
    v1, v2 = _valid(df1), _valid(df2)
    r1 = v1[v1[COL_DOI].isin(shared)].sort_values(COL_DOI).reset_index(drop=True)
    r2 = v2[v2[COL_DOI].isin(shared)].sort_values(COL_DOI).reset_index(drop=True)
    return r1, r2


def _multi_engine_dups(df):
    """DOIs cited by >1 distinct engine in the same run."""
    v = _valid(df)
    agg = (v.groupby(COL_DOI)[COL_AI_ENGINE]
             .agg(engine_count='nunique',
                  engines_list=lambda x: ', '.join(sorted(x.unique()))))
    multi_dois = agg[agg['engine_count'] > 1].index
    result = v[v[COL_DOI].isin(multi_dois)].copy()
    result = result.merge(agg, on=COL_DOI, how='left')
    return result.sort_values([COL_DOI, COL_AI_ENGINE]).reset_index(drop=True)


def _same_engine_cross_run(df1, df2):
    """Same engine cited the same DOI in both runs."""
    v1, v2 = _valid(df1).copy(), _valid(df2).copy()
    v1['_run'], v2['_run'] = 'Run 1', 'Run 2'
    keys1 = set(zip(v1[COL_AI_ENGINE], v1[COL_DOI]))
    keys2 = set(zip(v2[COL_AI_ENGINE], v2[COL_DOI]))
    shared = keys1 & keys2

    def _keep(d):
        idx = [i for i, (eng, doi) in
               enumerate(zip(d[COL_AI_ENGINE], d[COL_DOI]))
               if (eng, doi) in shared]
        return d.iloc[idx]

    combined = pd.concat([_keep(v1), _keep(v2)], ignore_index=True)
    return combined.sort_values([COL_AI_ENGINE, COL_DOI, '_run']).reset_index(drop=True)


# ══════════════════════════════════════════════════════════════════════════════
#  Low-level cell helpers
# ══════════════════════════════════════════════════════════════════════════════

def _cell(ws, r, c, value, font=None, fill=None, align=None, border=THIN):
    cell = ws.cell(row=r, column=c, value=value)
    if font:   cell.font      = font
    if fill:   cell.fill      = fill
    if align:  cell.alignment = align
    if border: cell.border    = border
    return cell


def _merge_title(ws, r, text, n_cols, size=13):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
    c = ws.cell(row=r, column=1, value=text)
    c.font = Font(name="Arial", bold=True, size=size, color="1F3864")
    c.alignment = CENTER
    ws.row_dimensions[r].height = 28


def _merge_info(ws, r, text, n_cols):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
    c = ws.cell(row=r, column=1, value=text)
    c.font = ITAL_FONT
    c.alignment = LEFT
    ws.row_dimensions[r].height = 16


def _header_row(ws, r, labels, fill):
    for c, lbl in enumerate(labels, 1):
        _cell(ws, r, c, lbl, font=WHITE_FONT, fill=fill, align=CENTER)
    ws.row_dimensions[r].height = 22


def _group_header(ws, r, text, n_cols, fill=GRP_FILL):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
    c = ws.cell(row=r, column=1, value=text)
    c.font = GRP_FONT; c.fill = fill; c.alignment = LEFT; c.border = THIN
    ws.row_dimensions[r].height = 16


def _engine_banner(ws, r, text, n_cols, fill):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
    c = ws.cell(row=r, column=1, value=text)
    c.font = ENG_FONT; c.fill = fill; c.alignment = LEFT; c.border = THIN
    ws.row_dimensions[r].height = 20


def _data_row(ws, r, values, left_cols, fill=None, row_h=30):
    for c, val in enumerate(values, 1):
        align = LEFT if c in left_cols else CENTER
        _cell(ws, r, c, val, font=NORM_FONT, fill=fill, align=align)
    ws.row_dimensions[r].height = row_h


def _col_w(ws, widths):
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w


def _new_sheet(writer, name):
    ws = writer.book.create_sheet(title=name)
    writer.sheets[name] = ws
    return ws


# ══════════════════════════════════════════════════════════════════════════════
#  Sheet writers
# ══════════════════════════════════════════════════════════════════════════════

def write_summary(writer, dup1, dup2, cross1, cross2,
                  multi1, multi2, same_eng, total1, total2):
    ws = _new_sheet(writer, "Summary")
    _merge_title(ws, 1, "Duplicate DOI Analysis — Summary", 4, size=14)
    ws.row_dimensions[1].height = 32
    _header_row(ws, 3, ["Metric", "Run 1 (0317)", "Run 2 (0319)", "Notes"], HEADER_FILL)

    se_dois = same_eng[COL_DOI].nunique()    if not same_eng.empty else 0
    se_r1   = same_eng[same_eng['_run'] == 'Run 1'].shape[0] if not same_eng.empty else 0
    se_r2   = same_eng[same_eng['_run'] == 'Run 2'].shape[0] if not same_eng.empty else 0
    m1_dois = multi1[COL_DOI].nunique()      if not multi1.empty else 0
    m2_dois = multi2[COL_DOI].nunique()      if not multi2.empty else 0

    rows = [
        ("── Within-run duplicates (any engine) ──", "", "", ""),
        ("Total citations loaded",              total1, total2,
         "All rows in each run file"),
        ("Duplicate DOIs (within run)",         dup1[COL_DOI].nunique(), dup2[COL_DOI].nunique(),
         "DOIs appearing >1× in the same run"),
        ("Affected citations (within run)",     len(dup1), len(dup2),
         "Total rows flagged"),
        ("── Cross-run duplicates (any engine) ──", "", "", ""),
        ("Shared DOIs across runs",             cross1[COL_DOI].nunique(), cross2[COL_DOI].nunique(),
         "Each side counted separately"),
        ("Affected citations — Run 1",          len(cross1), "—",
         "Run 1 rows with a cross-run DOI"),
        ("Affected citations — Run 2",          "—", len(cross2),
         "Run 2 rows with a cross-run DOI"),
        ("── Multi-engine duplicates (same day) ──", "", "", ""),
        ("DOIs cited by >1 engine",             m1_dois, m2_dois,
         "Same DOI, different engines, same run day"),
        ("Affected citations",                  len(multi1), len(multi2),
         "Total rows in multi-engine groups"),
        ("── Same-engine cross-run duplicates ──", "", "", ""),
        ("DOIs repeated by same engine",        se_dois, se_dois,
         "Engine cited same DOI in both runs"),
        ("Affected citations — Run 1",          se_r1, "—",
         "Run 1 rows in same-engine cross-run groups"),
        ("Affected citations — Run 2",          "—", se_r2,
         "Run 2 rows in same-engine cross-run groups"),
    ]

    section_idxs = {0, 4, 8, 11}
    for r_idx, (metric, v1, v2, note) in enumerate(rows):
        er = 4 + r_idx
        is_sec = r_idx in section_idxs
        fill = PatternFill("solid", fgColor="D9E1F2") if is_sec else (
               ALT_FILL if r_idx % 2 == 0 else None)
        font = Font(name="Arial", bold=True, size=10, color="1F3864") if is_sec else NORM_FONT
        for c_idx, val in enumerate([metric, v1, v2, note], 1):
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font      = BOLD_FONT if (c_idx == 1 and not is_sec) else font
            cell.alignment = LEFT if c_idx in (1, 4) else CENTER
            cell.border    = THIN
            if fill: cell.fill = fill
        ws.row_dimensions[er].height = 20 if is_sec else 18

    _col_w(ws, {'A': 42, 'B': 18, 'C': 18, 'D': 52})


def write_within_run(writer, sheet_name, df, hdr_fill, title):
    ws = _new_sheet(writer, sheet_name)
    n = len(OUTPUT_COLS)
    _merge_title(ws, 1, title, n)
    if df.empty:
        ws.cell(row=3, column=1, value="No duplicates found.").font = NORM_FONT
        return
    _merge_info(ws, 2,
                f"{df[COL_DOI].nunique()} duplicate DOI(s)  |  "
                f"{len(df)} affected citation(s)", n)
    _header_row(ws, 3, [c.replace('_', ' ').title() for c in OUTPUT_COLS], hdr_fill)
    left_cols = {OUTPUT_COLS.index(COL_DOI)+1, OUTPUT_COLS.index(COL_ORIG)+1}
    for i, (_, row) in enumerate(df.iterrows()):
        _data_row(ws, 4+i, [row[c] for c in OUTPUT_COLS],
                  left_cols, ALT_FILL if i % 2 else None)
    _col_w(ws, {'A': 18, 'B': 22, 'C': 40, 'D': 60})
    ws.freeze_panes = "A4"


def write_cross_run(writer, r1, r2):
    ws = _new_sheet(writer, "Cross_Run_Duplicates")
    n = len(OUTPUT_COLS) + 1
    left_cols = {OUTPUT_COLS.index(COL_DOI)+2, OUTPUT_COLS.index(COL_ORIG)+2}
    _merge_title(ws, 1,
                 "Cross-Run Duplicates  —  DOIs present in both Run 1 and Run 2", n)
    if r1.empty and r2.empty:
        ws.cell(row=3, column=1, value="No cross-run duplicates found.").font = NORM_FONT
        return
    shared = sorted(set(r1[COL_DOI]) | set(r2[COL_DOI]))
    _merge_info(ws, 2, f"{len(shared)} DOI(s) shared across both runs", n)
    _header_row(ws, 3,
                ["Run"] + [c.replace('_', ' ').title() for c in OUTPUT_COLS],
                CROSS_FILL)
    cur = 4
    for doi in shared:
        g1, g2 = r1[r1[COL_DOI]==doi], r2[r2[COL_DOI]==doi]
        _group_header(ws, cur, f"DOI: {doi}   ({len(g1)+len(g2)} citations)", n)
        cur += 1
        for _, row in g1.iterrows():
            _data_row(ws, cur, ["Run 1"]+[row[c] for c in OUTPUT_COLS], left_cols, RUN1_FILL)
            cur += 1
        for _, row in g2.iterrows():
            _data_row(ws, cur, ["Run 2"]+[row[c] for c in OUTPUT_COLS], left_cols, RUN2_FILL)
            cur += 1
        cur += 1
    _col_w(ws, {'A': 10, 'B': 18, 'C': 22, 'D': 40, 'E': 60})
    ws.freeze_panes = "A4"


def write_multi_engine(writer, sheet_name, df, title):
    """
    DOIs cited by more than one AI engine on the same run day.
    Extra columns prepended: '# Engines' and 'Engines' (comma-separated list).
    Grouped by DOI — each group header shows the engine count and names.
    """
    ws = _new_sheet(writer, sheet_name)
    extra = ["# Engines", "Engines"]
    cols  = extra + OUTPUT_COLS
    n     = len(cols)
    left_cols = {cols.index(COL_DOI)+1, cols.index(COL_ORIG)+1, cols.index("Engines")+1}

    _merge_title(ws, 1, title, n)
    if df.empty:
        ws.cell(row=3, column=1,
                value="No DOIs cited by multiple engines found.").font = NORM_FONT
        return

    _merge_info(ws, 2,
                f"{df[COL_DOI].nunique()} DOI(s) cited by >1 engine  |  "
                f"{len(df)} total citations", n)
    _header_row(ws, 3, [c.replace('_', ' ').title() for c in cols], MULTI_FILL)

    cur = 4
    for doi, grp in df.groupby(COL_DOI, sort=True):
        eng_count = int(grp['engine_count'].iloc[0])
        eng_list  = grp['engines_list'].iloc[0]
        _group_header(ws, cur,
                      f"DOI: {doi}   |   {eng_count} engine(s): {eng_list}"
                      f"   |   {len(grp)} citation(s)", n)
        cur += 1
        for i, (_, row) in enumerate(grp.iterrows()):
            vals = [eng_count, eng_list] + [row[c] for c in OUTPUT_COLS]
            _data_row(ws, cur, vals, left_cols, ALT_FILL if i % 2 else None)
            cur += 1
        cur += 1

    _col_w(ws, {'A': 12, 'B': 38, 'C': 20, 'D': 22, 'E': 40, 'F': 60})
    ws.freeze_panes = "A4"


def write_same_engine_cross_run(writer, df):
    """
    Same engine cited the same DOI in both runs.
    Grouped by engine (colour-banded banner), then by DOI.
    Run 1 rows in blue, Run 2 rows in green.
    """
    ws = _new_sheet(writer, "SameEngine_CrossRun")
    n = len(OUTPUT_COLS) + 1
    left_cols = {OUTPUT_COLS.index(COL_DOI)+2, OUTPUT_COLS.index(COL_ORIG)+2}

    _merge_title(ws, 1,
                 "Same-Engine Cross-Run Duplicates  —  "
                 "Same engine cited the same DOI in both Run 1 and Run 2", n)
    if df.empty:
        ws.cell(row=3, column=1,
                value="No same-engine cross-run duplicates found.").font = NORM_FONT
        return

    engines    = df[COL_AI_ENGINE].unique()
    total_dois = df.groupby(COL_AI_ENGINE)[COL_DOI].nunique().sum()
    _merge_info(ws, 2,
                f"{len(engines)} engine(s)  |  "
                f"{total_dois} unique DOI(s) cited in both runs  |  "
                f"{len(df)} total citations", n)
    _header_row(ws, 3,
                ["Run"] + [c.replace('_', ' ').title() for c in OUTPUT_COLS],
                SENG_FILL)

    cur = 4
    for eng_idx, (engine, eng_grp) in enumerate(df.groupby(COL_AI_ENGINE)):
        eng_fill = ENGINE_FILLS[eng_idx % len(ENGINE_FILLS)]
        eng_dois = eng_grp[COL_DOI].nunique()
        _engine_banner(ws, cur,
                       f"  Engine: {engine}   ({eng_dois} shared DOI(s))",
                       n, eng_fill)
        cur += 1

        for doi, doi_grp in eng_grp.groupby(COL_DOI):
            g1 = doi_grp[doi_grp['_run'] == 'Run 1']
            g2 = doi_grp[doi_grp['_run'] == 'Run 2']
            _group_header(ws, cur,
                          f"DOI: {doi}   "
                          f"({len(g1)} Run-1 citation(s),  {len(g2)} Run-2 citation(s))", n)
            cur += 1
            for _, row in g1.iterrows():
                _data_row(ws, cur, ["Run 1"]+[row[c] for c in OUTPUT_COLS],
                          left_cols, RUN1_FILL)
                cur += 1
            for _, row in g2.iterrows():
                _data_row(ws, cur, ["Run 2"]+[row[c] for c in OUTPUT_COLS],
                          left_cols, RUN2_FILL)
                cur += 1
            cur += 1
        cur += 1

    _col_w(ws, {'A': 10, 'B': 20, 'C': 22, 'D': 40, 'E': 60})
    ws.freeze_panes = "A4"


# ══════════════════════════════════════════════════════════════════════════════
#  Main
# ══════════════════════════════════════════════════════════════════════════════

def main():
    root = tk.Tk()
    root.withdraw()

    print("Select the Excel files for Run 1 (0317) and Run 2 (0319)...")
    file_paths = filedialog.askopenfilenames(
        title="Select Citation Lookup Results (both run files)",
        filetypes=[("Excel Files", "*.xlsx")]
    )
    if not file_paths:
        print("No files selected. Exiting.")
        return

    run_map = {}
    for fp in file_paths:
        detected = _detect_run(fp)
        if detected:
            run_map[fp] = detected
            print(f"  Detected Run {detected}: {os.path.basename(fp)}")
        else:
            ans = simpledialog.askstring(
                "Assign Run",
                f"Cannot auto-detect run for:\n{os.path.basename(fp)}\n\n"
                "Enter  1  (Run 1 / 0317)  or  2  (Run 2 / 0319):",
                parent=root
            )
            if ans in ('1', '2'):
                run_map[fp] = ans
                print(f"  Manually assigned Run {ans}: {os.path.basename(fp)}")
            else:
                print(f"  Skipping {os.path.basename(fp)}")

    run1_files = [fp for fp, r in run_map.items() if r == '1']
    run2_files = [fp for fp, r in run_map.items() if r == '2']

    if not run1_files:
        messagebox.showerror("Error", "No Run 1 (0317) file found. Exiting.")
        return
    if not run2_files:
        messagebox.showerror("Error", "No Run 2 (0319) file found. Exiting.")
        return

    def _load_many(fps):
        frames = [f for f in (_load_file(fp) for fp in fps) if f is not None]
        return pd.concat(frames, ignore_index=True) if frames \
               else pd.DataFrame(columns=OUTPUT_COLS)

    df1 = _load_many(run1_files)
    df2 = _load_many(run2_files)
    print(f"\n  Run 1 rows: {len(df1)}   Run 2 rows: {len(df2)}")

    dup1           = _within_run_dups(df1)
    dup2           = _within_run_dups(df2)
    cross1, cross2 = _cross_run_dups(df1, df2)
    multi1         = _multi_engine_dups(df1)
    multi2         = _multi_engine_dups(df2)
    same_eng       = _same_engine_cross_run(df1, df2)

    print(f"  Within-run dup DOIs    — Run 1: {dup1[COL_DOI].nunique()}  "
          f"Run 2: {dup2[COL_DOI].nunique()}")
    print(f"  Cross-run shared DOIs  : {cross1[COL_DOI].nunique()}")
    print(f"  Multi-engine DOIs      — Run 1: "
          f"{multi1[COL_DOI].nunique() if not multi1.empty else 0}  "
          f"Run 2: {multi2[COL_DOI].nunique() if not multi2.empty else 0}")
    print(f"  Same-engine cross-run  : "
          f"{same_eng[COL_DOI].nunique() if not same_eng.empty else 0} DOI(s)")

    out_dir  = os.path.dirname(file_paths[0])
    out_path = os.path.join(out_dir, "DOI_Duplicates.xlsx")

    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        if 'Sheet' in writer.book.sheetnames:
            del writer.book['Sheet']

        write_summary(writer,
                      dup1, dup2, cross1, cross2,
                      multi1, multi2, same_eng,
                      total1=len(df1), total2=len(df2))

        write_within_run(writer, "Run1_Duplicates", dup1, RUN1_FILL,
                         "Run 1 (0317) — Within-Run Duplicate DOIs")
        write_within_run(writer, "Run2_Duplicates", dup2, RUN2_FILL,
                         "Run 2 (0319) — Within-Run Duplicate DOIs")

        write_cross_run(writer, cross1, cross2)

        write_multi_engine(writer, "MultiEngine_Run1", multi1,
                           "Run 1 (0317) — DOIs Cited by Multiple AI Engines")
        write_multi_engine(writer, "MultiEngine_Run2", multi2,
                           "Run 2 (0319) — DOIs Cited by Multiple AI Engines")

        write_same_engine_cross_run(writer, same_eng)

    print(f"\n{'━'*65}")
    print(f"  Output -> {out_path}")
    print(f"{'━'*65}\n")


if __name__ == "__main__":
    main()